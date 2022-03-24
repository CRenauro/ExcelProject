
import openpyxl
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

s = Service("C:/Users/crenauro/Documents/selenium_training/chromedriver.exe")
driver = webdriver.Chrome(service=s)

driver.get("https://demoqa.com/automation-practice-form")
driver.maximize_window()

path = "C:/Users/crenauro/Documents/selenium_training/testdata1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
nrows = sheet_obj.max_row
ncol = sheet_obj.max_column
actions = ActionChains(driver)

print(nrows)
print(ncol)

for i in range(2, nrows+1):
    field_name = sheet_obj.cell(row=i, column=1).value
    xpath = sheet_obj.cell(row=i, column=2).value
    action_type = sheet_obj.cell(row=i, column=3).value
    data = sheet_obj.cell(row=i, column=4).value
    text = sheet_obj.cell(row=i, column=5).value

    if action_type == "TEXTBOX":
        try:
            print(xpath)
            driver.find_element(By.XPATH, xpath).send_keys(data)
            print("Entered Data")
            sheet_obj.cell(row=i, column=6).value = "PASS"
        except Exception:
            print("Unable to Enter Data")
            sheet_obj.cell(row=i, column=6).value = "FAIL"


    if action_type == "click":

        try:
            time.sleep(5)
            driver.find_element(By.XPATH, xpath).click()
            print("Clicked Successfully")
            sheet_obj.cell(row=i, column=6).value = "PASS"
        except Exception:
            print("Unable to Click")
            sheet_obj.cell(row=i, column=6).value = "FAIL"

    if action_type == "select":
        try:
            print(xpath)
            element = driver.find_element(By.XPATH, xpath)
            element.is_enabled()
            element.send_keys(Keys.CONTROL + "a")
            element.send_keys(text)
            from keyboard import press
            press('enter')
            print("Entered Data")
            sheet_obj.cell(row=i, column=6).value = "PASS"

        except Exception:
            print("Unable to Enter Data")
            sheet_obj.cell(row=i, column=6).value = "FAIL"




## full path causes permissions error but can't save results without it
wb_obj.save("C:/Users/crenauro/Documents/selenium_training/testdata1.xlsx")
